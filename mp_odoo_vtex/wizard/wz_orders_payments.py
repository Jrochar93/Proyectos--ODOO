# -*- coding: utf-8 -*-

from odoo import models, fields, api, _
import xlrd
from odoo.exceptions import RedirectWarning, UserError, ValidationError, AccessError
import openpyxl
import base64
from io import BytesIO
import logging
from unidecode import unidecode
import http.client
import json
import logging

_logger = logging.getLogger(__name__)


class WzRecalculateOrdersPaysVtex(models.TransientModel):
    _name = "wz.recalculate.orders.pays.vtex"
    _description = "Regenerar Pagos"

    start_date = fields.Date('Fecha de Inicio')
    end_date = fields.Date('Fecha de Inicio')

    type_process = fields.Selection([
        ('payment', 'Pagos'),
        ('payment_apply', 'Aplicar pago'),
        ('reconcile', 'Conciliar'),
    ], string='Tipo de Proceso')

    def generate_pays_to_order_all(self):
        vtex_orders = self.env['vtex.orders']
        item_payment_ids = self.env['vtex.orders.payments.items']
        payment_vtex_id_mp =  self.env['vtex.orders.payments']
        payment_vtex = self.env['vtex.orders.payments']
        vals_updt = {}
        tax_details_list = []
        list_d = []
        d1={}
        total_amount = total_amount_pay = total_commision_pay = total_shipping_pay = total_fee_pay =  total_fee_pay = cost_shipping_pay = am_f = am_iva = am_ica = total_chargue_shipp_pay = chargue_shipping_pay = 0

        res_orders = vtex_orders.search([
                                            '&',
                                            '&',
                                            ('create_date','>',self.start_date),
                                            ('create_date','<=',self.end_date),
                                            ('status','!=','canceled')
                                        ])
        # sync json
        for record_j in res_orders:
            record_j.sync_json_draft()
        

        for record in res_orders:
            if record.marketplaceOrderId:
                # validar si tiene pago generado
                get_data_pay = item_payment_ids.search([
                                                    '|',
                                                    '|',
                                                    '|',
                                                    ('id_ref_shipping_pay','ilike',str(record.name)),
                                                    ('id_ref_shipping_pay','ilike',str(record.shipment_id_cd)),
                                                    ('id_ref_shipping_pay','ilike',str(record.paymentIdMarketplace_cd)),
                                                    ('id_oper_payment_market','ilike',str(record.paymentIdMarketplace_cd)),
                                            ])
                
                if get_data_pay:
                    # buscar si tiene generado el pago
                    recorset_pay_vtex = payment_vtex_id_mp.search([('order_vtex_id','=',record.id)])
                    if not recorset_pay_vtex:
                        # crear pago
                        date_order_vtx = record.creationDate[:10]
                        vals_payment = {
                            'name': str(record.name),
                            'partner_id': record.partner_id.id,
                            'amount': float(record.value),
                            'date': date_order_vtx,
                            'order_vtex_id': record.id,
                            }
                        res_payment = payment_vtex.sudo().create(vals_payment)
                        res_payment_val = res_payment.id
                        vals_update = {
                                'payment_vtex_id_mp': res_payment_val,
                        }
                        write_order = record.sudo().write(vals_update)
                    else:
                        res_payment = recorset_pay_vtex

                for record_p in get_data_pay:
                    tax_details = record_p.tax_details_pay
                    if tax_details != '' or tax_details != [] or tax_details != '[]' or tax_details != '[]':
                        tax_details_f1 = tax_details.replace('[','')
                        tax_details_f2 = tax_details_f1.replace(']','')
                        tax_details_f3 = tax_details_f2.replace('"fuente,','"fuente",')
                        tax_details_f4 = tax_details_f3.replace('"iva,','"iva",')
                        tax_details_f5 = tax_details_f4.replace('"ica,','"ica",')
                        tax_details_f6_f1 = tax_details_f5.replace(', "detail"', '", "detail"')
                        tax_details_f6 = tax_details_f6_f1.replace(' ', '')

                        tax_det = tax_details_f6.split(sep='{', maxsplit=1)
                        tax_det = tax_details_f6.split(sep='}', maxsplit=2)
                        tax_det = tax_details_f6.split(sep='{', maxsplit=3)

                        if  tax_details != '['']':
                            if len(tax_det) == 4:
                                # todos los impuestos
                                # crear diccionarios
                                dict_1 = "{"+tax_det[1]
                                dict_2 = "{"+tax_det[2]
                                dict_3 = "{"+tax_det[3]

                                # replace extra
                                dict_1_f = dict_1.replace('},','}')
                                dict_2_f = dict_2.replace('},','}')
                                dict_3_f = dict_3.replace('},','}')

                                dict_1_f_1 = json.loads(dict_1_f)
                                dict_2_f_2 = json.loads(dict_2_f)
                                dict_3_f_2 = json.loads(dict_3_f)

                                if dict_1_f_1['financial_entity'] == 'fuente':
                                    am_f = abs(float(dict_1_f_1['amount']))
                                if dict_2_f_2['financial_entity'] == 'iva':
                                    am_iva = abs(float(dict_2_f_2['amount']))
                                if dict_3_f_2['financial_entity'] == 'ica':
                                    am_ica = abs(float(dict_2_f_2['amount']))
                            else:
                                # crear diccionarios
                                dict_1 = "{"+tax_det[1]
                                dict_2 = "{"+tax_det[2]

                                # replace extra
                                dict_1_f = dict_1.replace('},','}')
                                dict_2_f = dict_2.replace('},','}')

                                dict_1_f_1 = json.loads(dict_1_f)
                                dict_2_f_2 = json.loads(dict_2_f)

                                if dict_1_f_1['financial_entity'] == 'fuente':
                                    am_f = abs(float(dict_1_f_1['amount']))
                                if dict_2_f_2['financial_entity'] == 'iva':
                                    am_iva = abs(float(dict_2_f_2['amount']))

                        
                    # separar los valores 
                    if abs(float(record_p.value_sale_pay)) > 0:
                        total_amount = abs(float(record_p.value_sale_pay))
                    if abs(float(record_p.amount_net_oper_pay)) > 0:
                        total_amount_pay = abs(float(record_p.amount_net_oper_pay))
                    if abs(float(record_p.commision_meli_tax_pay)) > 0:
                        total_commision_pay = abs(float(record_p.commision_meli_tax_pay))
                    if  abs(float(record_p.cost_shipping_pay))  > 0:
                        cost_shipping_pay = abs(float(record_p.cost_shipping_pay))
                    if  abs(float(record_p.chargue_shipping_pay))  > 0:
                        chargue_shipping_pay = abs(float(record_p.chargue_shipping_pay))
                    #if float(record_p.total_fee_pay) > 0:
                    #    total_fee_pay = record_p.total_fee_pay

                    vals_updt = {
                        'total_amount': total_amount,
                        'total_amount_pay': total_amount_pay,
                        'total_commision_pay': total_commision_pay,
                        'total_shipping_pay': cost_shipping_pay,
                        'total_chargue_shipp_pay': chargue_shipping_pay,
                        'total_fee_pay': 0,
                        'total_tax_pay': am_iva,
                        'total_ica_pay': am_ica,
                        'total_rtefuente_pay': am_f,
                    }
                    res_payment.write(vals_updt)


class WzOpOrdersPaymentsVtex(models.TransientModel):
    _name = "wz.op.orders.payments.vtex"
    _description = "Operaciones multiples de pagos de ventas vtex"

    pays_ids_mp = fields.Many2many('vtex.orders.payments', 'wz_payment_order_vtex',string='Pagos ')
    pay_ids = fields.Many2many('vtex.orders.payments', 'wz_payment_order_vtex_account', string='Recibos de pago ')
    pays_to_reconcilie_ids = fields.Many2many('vtex.orders.payments', 'wz_pays_to_reconciliate_inv_vtex', string='Pagos ')

    # sync payment
    pays_sync_ids_mp = fields.Many2many('vtex.orders.payments', 'wz_payment_sync_order_vtex',string='Pagos ')

    @api.onchange('pays_ids_mp')
    def list_pays_ids(self):
        list_pays = []
        for record in self.env['vtex.orders.payments'].browse(self._context['active_ids']):
            if record.total_amount > 0 and not record.payment_id:
                list_pays.append(record.id)
            else:
                if record.payment_id:
                    raise ValidationError("El pago %s ya ha sido aplicado" % (record.name))
        
        self.pays_ids_mp = [(6,self,list_pays)]

    @api.onchange('pay_ids')
    def list_payment_ids(self):
        list_payment_ids = []
        for record in self.env['vtex.orders.payments'].browse(self._context['active_ids']):
            if record.payment_id.state == 'draft':
                list_payment_ids.append(record.id)
            else:
                if record.payment_id.state == 'posted':
                    raise ValidationError("El pago %s ya ha sido confirmado" % (record.name))
        
        self.pay_ids = [(6,self,list_payment_ids)]

    @api.onchange('pays_to_reconcilie_ids')
    def list_payment_to_reconciliate_ids(self):
        list_payment_ids = []
        for record in self.env['vtex.orders.payments'].browse(self._context['active_ids']):
            if record.payment_id.state == 'posted':
                list_payment_ids.append(record.id)
            else:
                if record.payment_id.state == 'draft':
                    raise ValidationError("El pago %s no ha sido confirmado" % (record.name))
        
        self.pays_to_reconcilie_ids = [(6,self,list_payment_ids)]


    @api.onchange('pays_sync_ids_mp')
    def list_pays_sync_ids_mp(self):
        list_payment_ids = []
        for record in self.env['vtex.orders.payments'].browse(self._context['active_ids']):
            if record:
                list_payment_ids.append(record.id)

        self.pays_sync_ids_mp = [(6,self,list_payment_ids)]


    def apply_pays_to_invoice(self):
        for record in self.pays_ids_mp:
            if record.total_amount > 0 and not record.payment_id:
                #aplicar pago a factura
                record.apply_payment()
            else:
                raise ValidationError("El pago %s ya ha sido aplicado" % (record.name))
            
    def confirm_payment_to_invoice(self):
        for record in self.pay_ids:
            if record.payment_id.state != 'posted':
                #confirmar pago a factura
                record.confirm_payment()
            else:
                raise ValidationError("El pago %s ya ha sido confirmado" % (record.name))

    def conciliate_payment_to_invoice(self):
        for record in self.pays_to_reconcilie_ids:
            if record.payment_id.state == 'posted':
                #conciltar pago con factura
                record.reconciliate_payment()
            else:
                raise ValidationError("El pago %s ya ha sido conciliado con la factura" % (record.name))

    def sync_multiple_payments(self):
        for record in self.pays_sync_ids_mp:
            if record:
                #conciltar pago con factura
                record.sync_payment()
           

            

class WzGetOrdersPaymentsVtex(models.TransientModel):
    _name = "wz.get.orders.payments.vtex"
    _description = "Pagos de ventas vtex"

    field_data = fields.Binary('Archivo Excel Pagos', required=True)
    field_data_shipping = fields.Binary('Archivo Excel Envios', required=True)
    name_field_data = fields.Char('Nombre archivo')
    name_field_data_ship = fields.Char('Nombre archivo Envio')

    format_import = fields.Selection([
        ('format_xlsx_1', 'Formato Meli Extendedido'),
        ('format_xlsx_2', 'Formato Meli normal')
    ], string='Formato de Excel',required=True)

    def upload_data(self):
        order_payment = self.env['vtex.orders.payments']
        order_payments_items = self.env['vtex.orders.payments.items']
        

        try:
            wb = openpyxl.load_workbook(filename=BytesIO(base64.b64decode(self.field_data)), read_only=True )
            sheet = wb.active
            
            # extemdido
            if self.format_import == 'format_xlsx_1':
                item_payment_ids = []
                for loadTemp in sheet.iter_rows(min_row=2, max_row=3000, min_col=1,max_col=42, values_only=True):
                    if(loadTemp[0] is not None):
                        # validar si existe el registro 
                        search_id_pay = self.env['vtex.orders.payments.items'].search([('name','=',str(loadTemp[0]))],limit=1)

                        if not search_id_pay.name or search_id_pay.name == '':
                            item_payment_ids.append((0,0,{
                                    'name': str(loadTemp[0]),
                                    'id_oper_payment_market': str(loadTemp[1]),
                                    'code_account_seller': str(loadTemp[2]),
                                    'type_medium_pay': str(loadTemp[3]),
                                    'medium_pay': str(loadTemp[4]),
                                    'country_origin_pay': str(loadTemp[5]),
                                    'value_sale_pay': str(loadTemp[7]),
                                    'currency_pay': str(loadTemp[8]),
                                    'date_sale_pay': str(loadTemp[10]),

                                    'commission_tax_pay': str(loadTemp[11]),
                                    'amount_oper_sale_pay': str(loadTemp[12]),
                                    'currency_settlement_pay': str(loadTemp[13]),
                                    'date_approve_pay': str(loadTemp[14]),
                                    'amount_net_oper_pay': str(loadTemp[15]),
                                    'data_extra': str(loadTemp[17]),


                                    'commision_meli_tax_pay': str(loadTemp[18]),
                                    'cost_shipping_pay': str(loadTemp[20]),
                                    'taxes_charged_pay': str(loadTemp[21]),

                                    # add more row

                                    'id_box_pay': str(loadTemp[24]),
                                    'id_branch_pay': str(loadTemp[25]),
                                    'name_branch_pay': str(loadTemp[26]),
                                    'id_box_defined_user_pay': str(loadTemp[27]),
                                    'name_branch_pay': str(loadTemp[28]),
                                    'id_branch_defined_user_pay': str(loadTemp[29]),

                                    

                                    'id_orden_meli_pay': str(loadTemp[30]),
                                    'id_ref_shipping_pay': str(loadTemp[31]),
                                    'mode_shipping_pay': str(loadTemp[32]),

                                    'number_inital_target_pay': str(loadTemp[38]),

                                    'id_ref_package_pay': str(loadTemp[33]),
                                    'tax_details_pay': str(loadTemp[34]),

                                    'sale_channel_pay': str(loadTemp[40]),
                                    'platform_payment_pay': str(loadTemp[41]),

                                }))
                            
                        id_oper = loadTemp[0]

                vals_payment = {
                    'name' : str(self.name_field_data),
                    'item_payment_ids' : item_payment_ids,
                }
                res_create_payment = order_payment.sudo().create(vals_payment)

            else:
                # normal
                if self.format_import == 'format_xlsx_2':
                    item_payment_ids = []
                    for loadTemp in sheet.iter_rows(min_row=2, max_row=3000, min_col=1,max_col=33, values_only=True):
                        if(loadTemp[0] is not None):
                            # validar si existe el registro 
                            search_id_pay = self.env['vtex.orders.payments.items'].search([('name','=',str(loadTemp[0]))],limit=1)

                            if not search_id_pay.name or search_id_pay.name == '':
                                item_payment_ids.append((0,0,{
                                        'name': str(loadTemp[0]),
                                        'id_oper_payment_market': str(loadTemp[1]),
                                        'code_account_seller': str(loadTemp[2]),
                                        'type_medium_pay': str(loadTemp[3]),
                                        'medium_pay': str(loadTemp[4]),
                                        'country_origin_pay': str(loadTemp[5]),
                                        'value_sale_pay': str(loadTemp[7]),
                                        'currency_pay': str(loadTemp[8]),
                                        'date_sale_pay': str(loadTemp[10]),

                                        'commission_tax_pay': str(loadTemp[11]),
                                        'amount_oper_sale_pay': str(loadTemp[12]),
                                        'currency_settlement_pay': str(loadTemp[13]),
                                        'date_approve_pay': str(loadTemp[14]),
                                        'amount_net_oper_pay': str(loadTemp[15]),
                                        'data_extra': str(loadTemp[17]),


                                        'commision_meli_tax_pay': str(loadTemp[18]),
                                        'cost_shipping_pay': str(loadTemp[20]),
                                        'taxes_charged_pay': str(loadTemp[21]),

                                        'id_orden_meli_pay': str(loadTemp[24]),
                                        'id_ref_shipping_pay': str(loadTemp[25]),
                                        'mode_shipping_pay': str(loadTemp[26]),

                                        'id_ref_package_pay': str(loadTemp[27]),
                                        'tax_details_pay': str(loadTemp[28]),
                                        'number_inital_target_pay': str(loadTemp[29]),

                                        'sale_channel_pay': str(loadTemp[31]),
                                        'platform_payment_pay': str(loadTemp[32]),
                                    }))
                    
                            id_oper = loadTemp[0]

                    vals_payment = {
                        'name' : str(self.name_field_data),
                        'item_payment_ids' : item_payment_ids,
                    }
                    res_create_payment = order_payment.sudo().create(vals_payment)

            # Envios
            
        
            try:
                wb2 = openpyxl.load_workbook(filename=BytesIO(base64.b64decode(self.field_data_shipping)),read_only=True)
                sheet_shipp = wb2.active
                item_pay_shipp_ids = []
                for loadTemp2 in sheet_shipp.iter_rows(min_row=2, max_row=500, min_col=1,max_col=20, values_only=True):
                    if(loadTemp2[0] is not None):
                        # validar si existe el registro 
                        search_id_pay = self.env['vtex.orders.payments.items'].search([
                                                                                    ('id_oper_payment_market','=',str(loadTemp2[12]))
                                                                                    ],limit=1)

                        if search_id_pay:
                            search_id_pay.write({'chargue_shipping_pay': str(loadTemp2[19])})

            except Exception as error2:
                raise UserError(_('Por favor, inserte un archivo válido 2%s' % error2))

        except Exception as error:
            raise UserError(_('Por favor, inserte un archivo válido %s' % error))
        
